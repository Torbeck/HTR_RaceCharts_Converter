"""Excel formatting utilities for HTR chart processing.

Applies formatting to Excel worksheets:
- Table creation with named styles
- Thin borders on all cells in a table range
- Auto-sized column widths based on content
- Frozen header row (pane at A2)

All three sheets are driven by per-table ExcelSettings from config.ini.
"""

from typing import List, Optional

import openpyxl
from openpyxl.styles import Border, Font, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.worksheet import Worksheet

from src.utils.ini_utils import ExcelSettings


def apply_table_formatting(
    ws: Worksheet,
    table_name: str,
    row_count: int,
    col_count: int,
    table_style: str,
    borders: bool,
    auto_size_columns: bool,
    freeze_header: bool,
) -> None:
    """Apply table formatting to a worksheet that already contains data.

    Table creation logic:
    - The table range spans from A1 to the last data cell.
    - The first row is treated as headers by the Table object.
    - The table_name must be unique within the workbook.

    Args:
        ws: Worksheet with headers in row 1 and data starting at row 2.
        table_name: Unique table name (e.g. 'race_data').
        row_count: Total number of rows including the header row.
        col_count: Number of columns.
        table_style: Excel table style name (e.g. 'TableStyleMedium9').
        borders: If True, apply thin borders to all cells in the table range.
        auto_size_columns: If True, compute and set column widths from content.
        freeze_header: If True, freeze the pane at A2 so the header row is fixed.
    """
    if row_count < 1 or col_count < 1:
        return

    last_col_letter = get_column_letter(col_count)
    table_ref = f"A1:{last_col_letter}{row_count}"

    # Create the Excel table with the specified style
    style_info = TableStyleInfo(
        name=table_style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    table = Table(displayName=table_name, ref=table_ref)
    table.tableStyleInfo = style_info
    ws.add_table(table)

    # Apply thin borders if requested
    if borders:
        _apply_thin_borders(ws, row_count, col_count)

    # Auto-size columns if requested
    if auto_size_columns:
        _auto_size_columns(ws, row_count, col_count)

    # Freeze header row if requested (pane at A2 keeps row 1 visible)
    if freeze_header:
        ws.freeze_panes = "A2"


def apply_sheet_formatting(
    ws: Worksheet,
    table_name: str,
    row_count: int,
    col_count: int,
    settings: ExcelSettings,
) -> None:
    """Apply INI-driven formatting to a worksheet.

    Args:
        ws: The worksheet to format.
        table_name: Unique Excel table name (e.g. 'race_data').
        row_count: Total rows including header.
        col_count: Number of columns.
        settings: ExcelSettings loaded from the corresponding INI section.
    """
    apply_table_formatting(
        ws=ws,
        table_name=table_name,
        row_count=row_count,
        col_count=col_count,
        table_style=settings.table_style,
        borders=settings.borders,
        auto_size_columns=settings.auto_size_columns,
        freeze_header=settings.freeze_header,
    )


def _apply_thin_borders(
    ws: Worksheet, row_count: int, col_count: int
) -> None:
    """Apply thin borders to every cell in the table range (A1 to last cell).

    Args:
        ws: Target worksheet.
        row_count: Total rows including header.
        col_count: Number of columns.
    """
    thin_side = Side(style="thin")
    thin_border = Border(
        left=thin_side, right=thin_side, top=thin_side, bottom=thin_side
    )
    for row in ws.iter_rows(min_row=1, max_row=row_count, max_col=col_count):
        for cell in row:
            cell.border = thin_border


def _auto_size_columns(
    ws: Worksheet, row_count: int, col_count: int
) -> None:
    """Compute and apply column widths based on cell content.

    Scans all cells in the table range and sets each column width to
    the length of the longest value plus a small padding. A minimum
    width of 8 and maximum of 50 are enforced to keep the sheet usable.

    Args:
        ws: Target worksheet.
        row_count: Total rows including header.
        col_count: Number of columns.
    """
    MIN_WIDTH = 8
    MAX_WIDTH = 50
    PADDING = 2

    for col_idx in range(1, col_count + 1):
        max_length = 0
        col_letter = get_column_letter(col_idx)
        for row_idx in range(1, row_count + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if cell.value is not None:
                cell_len = len(str(cell.value))
                if cell_len > max_length:
                    max_length = cell_len
        width = min(max(max_length + PADDING, MIN_WIDTH), MAX_WIDTH)
        ws.column_dimensions[col_letter].width = width
