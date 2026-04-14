"""Excel workbook utilities for HTR chart processing.

Builds the 3-sheet Excel workbook:
- Sheet 1 (Processed Race Data): formatted per [race_data] INI settings,
  with per-column number formats derived from fields.json field types.
- Sheet 2 (Points of Call by Distance): formatted per [points_call] INI settings.
- Sheet 3 (Fractional Times by Distance): formatted per [fractional_times] INI settings.
"""

import datetime
from typing import Any, Dict, List, Optional

import openpyxl

from src.utils.ini_utils import ExcelSettings
from src.utils.formatting_utils import apply_sheet_formatting


def build_workbook(
    processed_headers: List[str],
    processed_rows: List[List[str]],
    points_of_call_headers: List[str],
    points_of_call_rows: List[List[str]],
    fractional_times_headers: List[str],
    fractional_times_rows: List[List[str]],
    output_path: str,
    excel_settings: Optional[Dict[str, ExcelSettings]] = None,
    column_formats: Optional[List[Optional[str]]] = None,
) -> None:
    """Build and save an Excel workbook with three sheets.

    Sheet 1: Processed race data (headers + translated rows).
    Sheet 2: Points of Call by Distance.
    Sheet 3: Race Fractional Times by Distance.

    Each sheet's formatting is driven by its corresponding entry in
    excel_settings (keyed by 'race_data', 'points_call',
    'fractional_times').

    Args:
        processed_headers: Column headers for sheet 1.
        processed_rows: Data rows for sheet 1.
        points_of_call_headers: Column headers for sheet 2.
        points_of_call_rows: Data rows for sheet 2.
        fractional_times_headers: Column headers for sheet 3.
        fractional_times_rows: Data rows for sheet 3.
        output_path: Path to write the .xlsx file.
        excel_settings: Optional dict of per-table ExcelSettings from
            config.ini.  If provided, applies formatting to all sheets.
        column_formats: Optional list of Excel number-format strings for
            each column in sheet 1.  Aligned to processed_headers; a
            ``None`` entry means leave the column at Excel General format.
    """
    wb = openpyxl.Workbook()

    # ── Sheet 1: Processed Race Data ──────────────────────────────────
    ws1 = wb.active
    ws1.title = "Processed Race Data"
    _write_sheet(ws1, processed_headers, processed_rows, column_formats=column_formats)

    if column_formats:
        _apply_column_formats(ws1, column_formats, row_count=1 + len(processed_rows))

    if excel_settings and "race_data" in excel_settings:
        apply_sheet_formatting(
            ws1,
            table_name="race_data",
            row_count=1 + len(processed_rows),
            col_count=len(processed_headers),
            settings=excel_settings["race_data"],
        )

    # ── Sheet 2: Points of Call by Distance ───────────────────────────
    ws2 = wb.create_sheet(title="Points of Call by Distance")
    _write_sheet(ws2, points_of_call_headers, points_of_call_rows)

    if excel_settings and "points_call" in excel_settings:
        apply_sheet_formatting(
            ws2,
            table_name="points_call",
            row_count=1 + len(points_of_call_rows),
            col_count=len(points_of_call_headers),
            settings=excel_settings["points_call"],
        )

    # ── Sheet 3: Fractional Times by Distance ─────────────────────────
    ws3 = wb.create_sheet(title="Fractional Times by Distance")
    _write_sheet(ws3, fractional_times_headers, fractional_times_rows)

    if excel_settings and "fractional_times" in excel_settings:
        apply_sheet_formatting(
            ws3,
            table_name="fractional_times",
            row_count=1 + len(fractional_times_rows),
            col_count=len(fractional_times_headers),
            settings=excel_settings["fractional_times"],
        )

    from src.utils.file_utils import check_file_writable

    check_file_writable(output_path)
    wb.save(output_path)


def _write_sheet(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    headers: List[str],
    rows: List[List[str]],
    column_formats: Optional[List[Optional[str]]] = None,
) -> None:
    """Write headers and rows to a worksheet.

    When ``column_formats`` is provided, each cell value is coerced to the
    appropriate Python type (int, float, datetime) before writing so that
    Excel stores real numbers/dates rather than strings.  This prevents the
    "Number stored as text" green-triangle indicator in Excel.

    Args:
        ws: The openpyxl worksheet.
        headers: Column headers.
        rows: Data rows (typically all strings from CSV parsing).
        column_formats: Optional list of Excel format strings aligned to
            column positions (index 0 → column 1).  When supplied, each
            cell value is coerced before being written to the sheet.
    """
    ws.append(headers)
    for row in rows:
        if column_formats:
            coerced = [
                _coerce_cell_value(val, column_formats[i] if i < len(column_formats) else None)
                for i, val in enumerate(row)
            ]
            ws.append(coerced)
        else:
            ws.append(row)


def _coerce_cell_value(value: str, fmt: Optional[str]) -> Any:
    """Convert a raw string cell value to the appropriate Python type.

    Blank strings are always returned unchanged so that empty cells remain
    empty in the workbook.  Conversion failures fall back silently to the
    original string value so that no data is lost.

    Args:
        value: Raw string value from CSV parsing.
        fmt: Excel format string (e.g. ``"0"``, ``"$#,##0.00"``,
             ``"mm/dd/yyyy"``, ``"@"``), or ``None`` for General format.

    Returns:
        - ``int`` for Integer format (``"0"``)
        - ``float`` for Decimal (``"0.00"``) or Currency (``"$#,##0.00"``)
        - ``datetime.datetime`` for Date format (``"mm/dd/yyyy"``)
        - Original ``str`` for Text (``"@"``) or unknown/None formats,
          and as a fallback when conversion fails.
    """
    if not value:
        return value  # Preserve blanks / empty cells

    if fmt == "0":
        try:
            # Use int(float()) to handle "3" and "3.0" equally
            return int(float(value))
        except (ValueError, TypeError):
            return value

    if fmt in ("0.00", "$#,##0.00"):
        try:
            # Strip currency symbols and thousands separators before parsing
            cleaned = value.replace("$", "").replace(",", "").strip()
            return float(cleaned)
        except (ValueError, TypeError):
            return value

    if fmt == "mm/dd/yyyy":
        try:
            return datetime.datetime.strptime(value, "%m/%d/%Y")
        except (ValueError, TypeError):
            return value

    # Text ("@"), None (General), or any unrecognised format: keep as string
    return value


def _apply_column_formats(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    column_formats: List[Optional[str]],
    row_count: int,
) -> None:
    """Apply Excel number formats to data cells (rows 2 onward) in a worksheet.

    The header row (row 1) is intentionally skipped so that column headers
    remain as plain text regardless of the column format.

    Args:
        ws: Target worksheet.  Must already contain data written by
            :func:`_write_sheet`.
        column_formats: List of Excel format strings aligned to column
            positions (1-based: index 0 → column 1).  A ``None`` entry
            means leave the column at Excel General format (no-op).
        row_count: Total number of rows including the header row.
    """
    for col_idx, fmt in enumerate(column_formats, start=1):
        if fmt is None:
            continue
        for row_idx in range(2, row_count + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = fmt
