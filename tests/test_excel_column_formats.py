"""Tests for Excel column format enforcement.

Covers:
- get_column_formats: known types, unknown types, empty types, progress logging
- _coerce_cell_value: string → int/float/datetime/str type conversion
- _apply_column_formats: cell number_format set correctly for data rows
- build_workbook: column_formats applied to Sheet 1, not Sheet 2 or 3;
  values stored as real numbers/dates (not strings) to prevent Excel's
  "Number stored as text" green-triangle indicator.
- EXCEL_FORMATS sourced from src.schema_loader (single source of truth)
"""

import datetime
import os
import shutil
import tempfile
import unittest

import openpyxl

from src.schema_loader import EXCEL_FORMATS, VALID_TYPES, get_column_formats
from src.utils.excel_utils import _apply_column_formats, _coerce_cell_value, build_workbook


class TestGetColumnFormats(unittest.TestCase):
    """Verify get_column_formats returns correct format strings."""

    def _make_field(self, field_num, field_type):
        return {
            "field": field_num,
            "name": f"Field {field_num}",
            "type": field_type,
            "maxLength": "",
            "comments": "",
            "hasOptions": "",
        }

    def test_all_known_types_mapped(self):
        """Each VALID_TYPE should return its corresponding EXCEL_FORMATS value."""
        for type_name in VALID_TYPES:
            fields = [self._make_field(1, type_name)]
            result = get_column_formats(fields)
            self.assertEqual(result, [EXCEL_FORMATS[type_name]])

    def test_text_type(self):
        fields = [self._make_field(1, "Text")]
        self.assertEqual(get_column_formats(fields), ["@"])

    def test_integer_type(self):
        fields = [self._make_field(1, "Integer")]
        self.assertEqual(get_column_formats(fields), ["0"])

    def test_decimal_type(self):
        fields = [self._make_field(1, "Decimal")]
        self.assertEqual(get_column_formats(fields), ["0.00"])

    def test_date_type(self):
        fields = [self._make_field(1, "Date")]
        self.assertEqual(get_column_formats(fields), ["mm/dd/yyyy"])

    def test_currency_type(self):
        fields = [self._make_field(1, "Currency")]
        self.assertEqual(get_column_formats(fields), ["$#,##0.00"])

    def test_unknown_type_returns_none(self):
        """Unknown type should produce None (General format)."""
        fields = [self._make_field(1, "FooBar")]
        result = get_column_formats(fields)
        self.assertEqual(result, [None])

    def test_empty_type_returns_none(self):
        """Missing/empty type should produce None (General format)."""
        fields = [self._make_field(1, "")]
        result = get_column_formats(fields)
        self.assertEqual(result, [None])

    def test_multiple_fields_mixed_types(self):
        """Multiple fields produce one entry per field."""
        fields = [
            self._make_field(1, "Text"),
            self._make_field(2, "Date"),
            self._make_field(3, ""),
            self._make_field(4, "Integer"),
        ]
        result = get_column_formats(fields)
        self.assertEqual(result, ["@", "mm/dd/yyyy", None, "0"])

    def test_length_matches_schema_length(self):
        """Return list length must match number of fields."""
        fields = [self._make_field(i + 1, "Text") for i in range(10)]
        result = get_column_formats(fields)
        self.assertEqual(len(result), 10)

    def test_progress_called_for_unknown_type(self):
        """progress callback should be called for unmapped types."""
        warnings = []
        fields = [self._make_field(7, "Unknown")]
        get_column_formats(fields, progress=warnings.append)
        self.assertTrue(any("WARNING" in w for w in warnings))
        self.assertTrue(any("Field 7" in w for w in warnings))

    def test_progress_not_called_for_known_type(self):
        """progress callback should NOT be called for mapped types."""
        warnings = []
        fields = [self._make_field(1, "Text")]
        get_column_formats(fields, progress=warnings.append)
        self.assertEqual(warnings, [])

    def test_progress_none_safe(self):
        """Passing progress=None should not raise even for unmapped types."""
        fields = [self._make_field(1, "Unknown")]
        try:
            get_column_formats(fields, progress=None)
        except Exception as e:
            self.fail(f"get_column_formats raised unexpectedly: {e}")


class TestApplyColumnFormats(unittest.TestCase):
    """Verify _apply_column_formats sets cell.number_format correctly."""

    def _make_ws(self, headers, rows):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(headers)
        for row in rows:
            ws.append(row)
        return ws

    def test_format_applied_to_data_rows(self):
        ws = self._make_ws(["A", "B"], [["1", "2"], ["3", "4"]])
        _apply_column_formats(ws, ["0", "@"], row_count=3)
        # Column 1: Integer format "0"
        self.assertEqual(ws.cell(row=2, column=1).number_format, "0")
        self.assertEqual(ws.cell(row=3, column=1).number_format, "0")
        # Column 2: Text format "@"
        self.assertEqual(ws.cell(row=2, column=2).number_format, "@")
        self.assertEqual(ws.cell(row=3, column=2).number_format, "@")

    def test_header_row_not_formatted(self):
        """Row 1 (header) must not have its number_format changed."""
        ws = self._make_ws(["MyHeader"], [["value"]])
        _apply_column_formats(ws, ["0"], row_count=2)
        # The default number format for an unset cell is "General"
        header_fmt = ws.cell(row=1, column=1).number_format
        self.assertNotEqual(header_fmt, "0")

    def test_none_format_skipped(self):
        """None entries must leave cells at their default format."""
        ws = self._make_ws(["A", "B"], [["1", "2"]])
        _apply_column_formats(ws, [None, "mm/dd/yyyy"], row_count=2)
        # Column 1: None → untouched (openpyxl default is "General")
        self.assertNotEqual(ws.cell(row=2, column=1).number_format, "mm/dd/yyyy")
        # Column 2: Date format applied
        self.assertEqual(ws.cell(row=2, column=2).number_format, "mm/dd/yyyy")

    def test_empty_column_formats_list(self):
        """Empty column_formats list should apply nothing without error."""
        ws = self._make_ws(["A"], [["val"]])
        try:
            _apply_column_formats(ws, [], row_count=2)
        except Exception as e:
            self.fail(f"_apply_column_formats raised unexpectedly: {e}")

    def test_header_only_sheet_no_error(self):
        """row_count=1 (header only, no data rows) should be a no-op."""
        ws = self._make_ws(["A"], [])
        try:
            _apply_column_formats(ws, ["0"], row_count=1)
        except Exception as e:
            self.fail(f"_apply_column_formats raised unexpectedly: {e}")


class TestCellValueCoercion(unittest.TestCase):
    """Unit tests for _coerce_cell_value type conversion."""

    def test_integer_format_converts_to_int(self):
        result = _coerce_cell_value("42", "0")
        self.assertEqual(result, 42)
        self.assertIsInstance(result, int)

    def test_integer_format_whole_float_string(self):
        """String "3.0" should be converted to int 3 for Integer format."""
        result = _coerce_cell_value("3.0", "0")
        self.assertEqual(result, 3)
        self.assertIsInstance(result, int)

    def test_decimal_format_converts_to_float(self):
        result = _coerce_cell_value("3.14", "0.00")
        self.assertAlmostEqual(result, 3.14)
        self.assertIsInstance(result, float)

    def test_currency_format_converts_to_float(self):
        result = _coerce_cell_value("125000.00", "$#,##0.00")
        self.assertAlmostEqual(result, 125000.0)
        self.assertIsInstance(result, float)

    def test_currency_format_strips_symbols(self):
        """Values with '$' and ',' should be parsed correctly."""
        result = _coerce_cell_value("$1,234.56", "$#,##0.00")
        self.assertAlmostEqual(result, 1234.56)
        self.assertIsInstance(result, float)

    def test_currency_format_no_decimals(self):
        result = _coerce_cell_value("125000", "$#,##0.00")
        self.assertAlmostEqual(result, 125000.0)
        self.assertIsInstance(result, float)

    def test_date_format_converts_to_datetime(self):
        result = _coerce_cell_value("04/05/2024", "mm/dd/yyyy")
        self.assertEqual(result, datetime.datetime(2024, 4, 5))
        self.assertIsInstance(result, datetime.datetime)

    def test_text_format_keeps_string(self):
        result = _coerce_cell_value("hello", "@")
        self.assertEqual(result, "hello")
        self.assertIsInstance(result, str)

    def test_none_format_keeps_string(self):
        result = _coerce_cell_value("hello", None)
        self.assertEqual(result, "hello")
        self.assertIsInstance(result, str)

    def test_blank_always_returned_unchanged_for_all_formats(self):
        """Blank strings must remain blank for every format to preserve empty cells."""
        for fmt in ["0", "0.00", "$#,##0.00", "mm/dd/yyyy", "@", None]:
            with self.subTest(fmt=fmt):
                result = _coerce_cell_value("", fmt)
                self.assertEqual(result, "")

    def test_invalid_integer_falls_back_to_string(self):
        result = _coerce_cell_value("not_a_number", "0")
        self.assertEqual(result, "not_a_number")
        self.assertIsInstance(result, str)

    def test_invalid_float_falls_back_to_string(self):
        result = _coerce_cell_value("abc", "0.00")
        self.assertEqual(result, "abc")
        self.assertIsInstance(result, str)

    def test_invalid_date_falls_back_to_string(self):
        result = _coerce_cell_value("not_a_date", "mm/dd/yyyy")
        self.assertEqual(result, "not_a_date")
        self.assertIsInstance(result, str)

    def test_unknown_format_keeps_string(self):
        """Any unrecognised format string should return the value unchanged."""
        result = _coerce_cell_value("42", "##0.0E+0")
        self.assertEqual(result, "42")
        self.assertIsInstance(result, str)


class TestBuildWorkbookColumnFormats(unittest.TestCase):
    """Verify build_workbook applies column_formats to Sheet 1 data cells."""

    def setUp(self):
        self._tmpdir = tempfile.mkdtemp()
        self._output_path = os.path.join(self._tmpdir, "test_output.xlsx")

    def tearDown(self):
        shutil.rmtree(self._tmpdir, ignore_errors=True)

    def _build(self, headers, rows, col_fmts=None):
        build_workbook(
            processed_headers=headers,
            processed_rows=rows,
            points_of_call_headers=["dist", "c1"],
            points_of_call_rows=[["550", "1"]],
            fractional_times_headers=["dist", "f1"],
            fractional_times_rows=[["550", "22"]],
            output_path=self._output_path,
            column_formats=col_fmts,
        )
        return openpyxl.load_workbook(self._output_path)

    def test_column_formats_applied_sheet1(self):
        """Column formats from fields.json types set on Sheet 1 data cells."""
        headers = ["Track", "Date", "Race Number"]
        rows = [["CD", "04/05/2024", "3"]]
        # Text, Date, Integer
        col_fmts = ["@", "mm/dd/yyyy", "0"]
        wb = self._build(headers, rows, col_fmts)
        ws = wb.active  # Sheet 1
        self.assertEqual(ws.cell(row=2, column=1).number_format, "@")
        self.assertEqual(ws.cell(row=2, column=2).number_format, "mm/dd/yyyy")
        self.assertEqual(ws.cell(row=2, column=3).number_format, "0")

    def test_no_column_formats_no_error(self):
        """build_workbook without column_formats should not raise."""
        headers = ["A", "B"]
        rows = [["x", "y"]]
        try:
            self._build(headers, rows, col_fmts=None)
        except Exception as e:
            self.fail(f"build_workbook raised unexpectedly: {e}")

    def test_header_row_not_formatted_sheet1(self):
        """Header row (row 1) in Sheet 1 must not be number-formatted."""
        headers = ["Track"]
        rows = [["CD"]]
        col_fmts = ["@"]
        wb = self._build(headers, rows, col_fmts)
        ws = wb.active
        # openpyxl default for unset cells is "General"
        self.assertNotEqual(ws.cell(row=1, column=1).number_format, "@")

    def test_sheet2_unaffected_by_column_formats(self):
        """Sheet 2 (Points of Call) cells must not receive column_formats."""
        headers = ["Track"]
        rows = [["CD"]]
        col_fmts = ["@"]
        wb = self._build(headers, rows, col_fmts)
        ws2 = wb.worksheets[1]
        # "dist" header and data should be at General (not "@")
        self.assertNotEqual(ws2.cell(row=2, column=1).number_format, "@")

    def test_none_formats_leave_cells_general(self):
        """None entries in column_formats leave cells at General format."""
        headers = ["A", "B"]
        rows = [["x", "42"]]
        col_fmts = [None, "0"]
        wb = self._build(headers, rows, col_fmts)
        ws = wb.active
        self.assertNotEqual(ws.cell(row=2, column=1).number_format, "0")
        self.assertEqual(ws.cell(row=2, column=2).number_format, "0")

    # ── Value-type tests: ensure real numbers/dates are written (no green triangle) ──

    def test_integer_column_value_is_numeric(self):
        """Integer column must store an int value, not a string."""
        headers = ["Race Number"]
        rows = [["42"]]
        col_fmts = ["0"]
        wb = self._build(headers, rows, col_fmts)
        ws = wb.active
        cell_value = ws.cell(row=2, column=1).value
        self.assertNotIsInstance(
            cell_value, str, "Integer cell must not store a string value"
        )
        self.assertIsInstance(cell_value, (int, float))
        self.assertEqual(cell_value, 42)

    def test_currency_column_value_is_numeric(self):
        """Currency column must store a float value, not a string."""
        headers = ["Purse"]
        rows = [["125000.00"]]
        col_fmts = ["$#,##0.00"]
        wb = self._build(headers, rows, col_fmts)
        ws = wb.active
        cell_value = ws.cell(row=2, column=1).value
        self.assertNotIsInstance(
            cell_value, str, "Currency cell must not store a string value"
        )
        self.assertIsInstance(cell_value, (int, float))
        self.assertAlmostEqual(cell_value, 125000.0)

    def test_decimal_column_value_is_numeric(self):
        """Decimal column must store a float value, not a string."""
        headers = ["Odds"]
        rows = [["5.50"]]
        col_fmts = ["0.00"]
        wb = self._build(headers, rows, col_fmts)
        ws = wb.active
        cell_value = ws.cell(row=2, column=1).value
        self.assertNotIsInstance(
            cell_value, str, "Decimal cell must not store a string value"
        )
        self.assertIsInstance(cell_value, (int, float))
        self.assertAlmostEqual(cell_value, 5.5)

    def test_date_column_value_is_datetime(self):
        """Date column must store a datetime value, not a string."""
        headers = ["Race Date"]
        rows = [["04/05/2024"]]
        col_fmts = ["mm/dd/yyyy"]
        wb = self._build(headers, rows, col_fmts)
        ws = wb.active
        cell_value = ws.cell(row=2, column=1).value
        self.assertNotIsInstance(
            cell_value, str, "Date cell must not store a string value"
        )
        self.assertIsInstance(cell_value, datetime.datetime)
        self.assertEqual(cell_value, datetime.datetime(2024, 4, 5))

    def test_text_column_value_stays_string(self):
        """Text column values must remain strings."""
        headers = ["Track"]
        rows = [["CD"]]
        col_fmts = ["@"]
        wb = self._build(headers, rows, col_fmts)
        ws = wb.active
        cell_value = ws.cell(row=2, column=1).value
        self.assertIsInstance(cell_value, str)
        self.assertEqual(cell_value, "CD")

    def test_blank_cell_stays_blank_for_numeric_format(self):
        """Blank string values must remain blank (None) even for numeric columns."""
        headers = ["Race Number"]
        rows = [["", "42"]]
        col_fmts = ["0"]
        wb = self._build(headers, rows, col_fmts)
        ws = wb.active
        # Empty string written → cell should be None or empty string, not a number
        blank_value = ws.cell(row=2, column=1).value
        self.assertFalse(
            blank_value,
            "Blank cell in numeric column must remain blank/None, not be converted",
        )

    def test_integer_column_number_format_correct_after_coercion(self):
        """number_format must still be '0' for Integer columns after value coercion."""
        headers = ["Race Number"]
        rows = [["7"]]
        col_fmts = ["0"]
        wb = self._build(headers, rows, col_fmts)
        ws = wb.active
        self.assertEqual(ws.cell(row=2, column=1).number_format, "0")

    def test_currency_column_number_format_correct_after_coercion(self):
        """number_format must still be '$#,##0.00' after value coercion."""
        headers = ["Purse"]
        rows = [["50000.00"]]
        col_fmts = ["$#,##0.00"]
        wb = self._build(headers, rows, col_fmts)
        ws = wb.active
        self.assertEqual(ws.cell(row=2, column=1).number_format, "$#,##0.00")


class TestExcelFormatsSourcedFromSchemaLoader(unittest.TestCase):
    """Verify EXCEL_FORMATS and VALID_TYPES come from src.schema_loader."""

    def test_excel_formats_importable_from_schema_loader(self):
        from src.schema_loader import EXCEL_FORMATS as EF
        self.assertIsInstance(EF, dict)
        self.assertIn("Text", EF)
        self.assertIn("Date", EF)

    def test_valid_types_importable_from_schema_loader(self):
        from src.schema_loader import VALID_TYPES as VT
        self.assertIsInstance(VT, list)
        self.assertEqual(set(VT), set(EXCEL_FORMATS.keys()))

    def test_schema_editor_uses_same_excel_formats(self):
        """tools.schema_editor must re-export the same EXCEL_FORMATS object."""
        from tools.schema_editor import EXCEL_FORMATS as SE_EF
        from src.schema_loader import EXCEL_FORMATS as SL_EF
        self.assertIs(SE_EF, SL_EF)

    def test_schema_editor_uses_same_valid_types(self):
        from tools.schema_editor import VALID_TYPES as SE_VT
        from src.schema_loader import VALID_TYPES as SL_VT
        self.assertIs(SE_VT, SL_VT)


if __name__ == "__main__":
    unittest.main()
